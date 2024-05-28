import openpyxl as xlx
import xlrd

class xlFile:
    '''
    Convenience class for supporting .xlsx and .xls files
    without caring about details.
    '''

    def __init__(self, filename):
        self.wb = None

    @staticmethod
    def load_excel_file(filename):
        try:                    # Most of the time we will get .xlsx files, so let's try that
            return xlsxFile(filename)
        except xlx.utils.exceptions.InvalidFileException:
            # Try open an .xls file
            return xlsFile(filename) # Need exception handling here too?


class xlsxFile:
    def __init__(self, filename):
        self.wb = xlx.load_workbook(filename, data_only=True)

    def sheetnames(self):
        return self.wb.sheetnames


    def sheet(self, sheetname):
        return self.wb[sheetname]

        
class xlsFile:
    def __init__(self, filename):
        self.wb = xlrd.open_workbook(filename)

    def sheetnames(self):
        return self.wb.sheetnames()
